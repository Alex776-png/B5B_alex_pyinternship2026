import tkinter as tk
from tkinter import messagebox
from fpdf import FPDF
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import sys


class InvoiceGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("Simple Invoice Generator")
        self.root.geometry("480x620")
        self.root.resizable(False, False)

        # Title
        title = tk.Label(root, text="INVOICE GENERATOR", font=("Arial", 18, "bold"), fg="#1a5276")
        title.pack(pady=15)

        # Frame for all input fields
        form = tk.Frame(root)
        form.pack(pady=5)

        # All labels and entry boxes
        self.fields = {}
        labels = [
            "Client Name",
            "Address",
            "Item Description",
            "Quantity",
            "Rate per Day",
            "Customs Ref",
            "HS Code",
            "Country",
            "Status"
        ]

        for i, text in enumerate(labels):
            tk.Label(form, text=text + ":", font=("Arial", 10), anchor="w", width=18).grid(row=i, column=0, pady=6, sticky="w")
            entry = tk.Entry(form, width=32, font=("Arial", 10))
            entry.grid(row=i, column=1, pady=6, padx=5)
            self.fields[text] = entry

        # Buttons Frame
        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=25)

        tk.Button(btn_frame, text="Invoice", width=12, bg="#27ae60", fg="white",
                  font=("Arial", 11, "bold"), command=self.generate_invoice).grid(row=0, column=0, padx=8)

        tk.Button(btn_frame, text="Reset", width=12, bg="#f39c12", fg="white",
                  font=("Arial", 11, "bold"), command=self.reset_fields).grid(row=0, column=1, padx=8)

        tk.Button(btn_frame, text="Exit", width=12, bg="#c0392b", fg="white",
                  font=("Arial", 11, "bold"), command=self.root.destroy).grid(row=0, column=2, padx=8)

    def get_data(self):
        """Collect all values from the entry boxes"""
        data = {}
        for key, entry in self.fields.items():
            data[key] = entry.get().strip()
        return data

    def generate_invoice(self):
        data = self.get_data()

        # Basic validation
        if not data["Client Name"] or not data["Item Description"]:
            messagebox.showerror("Error", "Client Name and Item Description are required!")
            return

        try:
            qty = float(data["Quantity"])
            rate = float(data["Rate per Day"])
            total = qty * rate
        except ValueError:
            messagebox.showerror("Error", "Quantity and Rate per Day must be numbers!")
            return

        # Create Invoice Number & Date
        now = datetime.now()
        invoice_no = "INV-" + now.strftime("%Y%m%d%H%M%S")
        invoice_date = now.strftime("%d-%m-%Y")

        # ---------- Create PDF ----------
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 20)
        pdf.cell(0, 15, "INVOICE", ln=True, align="C")
        pdf.ln(5)

        pdf.set_font("Arial", size=11)
        pdf.cell(0, 8, f"Invoice No : {invoice_no}", ln=True)
        pdf.cell(0, 8, f"Date       : {invoice_date}", ln=True)
        pdf.ln(8)

        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, "Client Details", ln=True)
        pdf.set_font("Arial", size=11)
        pdf.cell(0, 7, f"Name    : {data['Client Name']}", ln=True)
        pdf.multi_cell(0, 7, f"Address : {data['Address']}")
        pdf.ln(5)

        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, "Item Details", ln=True)
        pdf.set_font("Arial", size=11)
        pdf.cell(0, 7, f"Description : {data['Item Description']}", ln=True)
        pdf.cell(0, 7, f"Quantity    : {data['Quantity']}", ln=True)
        pdf.cell(0, 7, f"Rate/Day    : {data['Rate per Day']}", ln=True)
        pdf.cell(0, 7, f"Total       : {total:.2f}", ln=True)
        pdf.ln(5)

        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, "Customs & Shipping Info", ln=True)
        pdf.set_font("Arial", size=11)
        pdf.cell(0, 7, f"Customs Ref : {data['Customs Ref']}", ln=True)
        pdf.cell(0, 7, f"HS Code     : {data['HS Code']}", ln=True)
        pdf.cell(0, 7, f"Country     : {data['Country']}", ln=True)
        pdf.cell(0, 7, f"Status      : {data['Status']}", ln=True)

        pdf.ln(15)
        pdf.set_font("Arial", "I", 10)
        pdf.cell(0, 8, "Thank you for your business!", ln=True, align="C")

        # Save PDF
        pdf_filename = f"Invoice_{invoice_no}.pdf"
        pdf.output(pdf_filename)

        # ---------- Save to Excel ----------
        excel_file = "Invoices_Data.xlsx"

        if not os.path.exists(excel_file):
            # Create new workbook if file does not exist
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoices"
            headers = ["Invoice No", "Date", "Client Name", "Address", "Item Description",
                       "Quantity", "Rate per Day", "Total", "Customs Ref", "HS Code",
                       "Country", "Status"]
            ws.append(headers)
        else:
            wb = load_workbook(excel_file)
            ws = wb.active

        # Add the new row
        row = [
            invoice_no,
            invoice_date,
            data["Client Name"],
            data["Address"],
            data["Item Description"],
            data["Quantity"],
            data["Rate per Day"],
            total,
            data["Customs Ref"],
            data["HS Code"],
            data["Country"],
            data["Status"]
        ]
        ws.append(row)
        wb.save(excel_file)

        messagebox.showinfo("Success",
                            f"Invoice generated successfully!\n\n"
                            f"PDF  → {pdf_filename}\n"
                            f"Excel → {excel_file}")

    def reset_fields(self):
        """Clear all entry boxes"""
        for entry in self.fields.values():
            entry.delete(0, tk.END)


# ---------- Main ----------
if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceGenerator(root)
    root.mainloop()